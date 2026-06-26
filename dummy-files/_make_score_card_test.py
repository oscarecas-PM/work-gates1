"""Generate synthetic test data for Employee Score Card.

Designs the test scenarios from the implementation plan:
- Two xlsx files with overlapping wildcards (dedup test)
- Multi-week incremental charges (Bob on WO-400: 5, 7, 8 per week)
- Split operation flagged (WO-100/0010, Alice + Bob, 60/40 of charges)
- ACTIVE op with no Hours Earned row (WO-100/0020) -- unmatched
- Whole new WO with no Hours Earned (WO-300) -- surfaces in unmatched report at top

v1.7 additions:
- Each charge row now carries optional CCC and Work Center columns.
- Charlie (B003) is tagged as a quality inspector (CCC=QA, WC=WC-INSP) so the
  CCC filter chip removes Charlie with one click.
- A new tech, Diana Park (B004), is tagged as the cleaning crew: CCC=TECH (same
  as Alice/Bob) but Work Center = WC-CLEAN (different). The Work Center filter
  alone, or composed with CCC=QA, exercises the exclusion logic against the
  user's real-world scenario:
    - Exclude CCC = QA           -> hides Charlie only.
    - Exclude WC = WC-CLEAN       -> hides Diana only.
    - Exclude both at once        -> only Alice + Bob remain (the "real" techs).
    - Click Charlie's chip while  -> manual override re-includes Charlie even
      QA is excluded                 though the auto rule wants to hide them.
"""
from openpyxl import Workbook
from datetime import datetime
from pathlib import Path

OUT = Path(__file__).parent

HEADERS = ["Order No", "Week Ending", "Employee Name", "Badge", "Part No.",
           "Part Description", "Operation", "Total Hours", "Status",
           "CCC", "Work Center"]

def make_xlsx(filename, rows, export_label):
    """Write an .xlsx with a 3-row export-metadata block, then headers on row 4, then data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Charges"
    ws["A1"] = f"Charge Export -- {export_label}"
    ws["A2"] = f"Generated: {datetime.now().isoformat()}"
    ws["A3"] = "Source: synthetic test data"
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=4, column=c, value=h)
    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(OUT / filename)
    print(f"Wrote {filename} ({len(rows)} data rows)")

# All weeks formatted MM/DD/YYYY HH:MM:SS AM/PM as strings
def w(month, day, year=2026):
    return f"{month}/{day}/{year} 12:00:00 AM"

# ----------------------------------------------------------------------------
# Per-tech CCC and Work Center tags (v1.7)
#   B001 Alice:   TECH / WC-MFG-1   (real technician)
#   B002 Bob:     TECH / WC-MFG-1   (real technician)
#   B003 Charlie: QA   / WC-INSP    (quality inspector -- excluded via CCC)
#   B004 Diana:   TECH / WC-CLEAN   (cleaning crew -- same CCC as techs,
#                                    different Work Center)
# ----------------------------------------------------------------------------
TECH_CCC, TECH_WC = "TECH", "WC-MFG-1"
QA_CCC,   QA_WC   = "QA",   "WC-INSP"
CLEAN_CCC, CLEAN_WC = "TECH", "WC-CLEAN"

# ----------------------------------------------------------------------------
# Scenario coverage (unchanged from earlier versions, plus Diana's new WO-600):
# Scenario 1: WO-100, Op 0010, "Wing Spar Machining", part 555-01 (SPLIT)
#   Alice incremental w1=10, w2=20 (charges 30 total). Bob w2=20.
#   Closes Mar 2026, Hours Earned = 60. -> Alice 36, Bob 24.
# Scenario 2: WO-100, Op 0020 -- ACTIVE (no Hours Earned row). Alice 5h.
# Scenario 3: WO-200, Op 0010, "Fuselage Drill", part 777-12. Charlie 50h.
#   Closes Apr 2026, Hours Earned = 40. CPI 0.8.
# Scenario 4: WO-300, Op 0030, "Inspect", part 999-05. Alice 10h, Charlie 8h.
#   ACTIVE forever. Unmatched.
# Scenario 5: WO-400, Op 0010, "Sheet Metal Cut", part 555-02. Bob multi-week
#   incremental 5/7/8 (20 total). Closes Mar 2026, Hours Earned 25. CPI 1.25.
# Scenario 6: WO-500, Op 0010, "Test", part 555-01. Bob 30h. Closes Apr 2026,
#   Hours Earned 25. CPI 0.83.
# Scenario 7 (v1.7): WO-600, Op 0010, "Cleaning Bay 5", part 555-03.
#   Diana 8h. Closes Apr 2026, Hours Earned 8. CPI 1.0. Tags Diana as CLEAN.
# ----------------------------------------------------------------------------

# File A: covers parts 555-* and 777-*
file_a_rows = [
    # WO-100/0010 split op (incremental: Alice 10 + 20 = 30 total, Bob 20)
    ["WO-100", w(2, 28), "Alice Johnson", "B001", "555-01", "Wing Spar Machining", "0010", 10.0, "ACTIVE", TECH_CCC, TECH_WC],
    ["WO-100", w(3, 7),  "Alice Johnson", "B001", "555-01", "Wing Spar Machining", "0010", 20.0, "CLOSE",  TECH_CCC, TECH_WC],
    ["WO-100", w(3, 7),  "Bob Lee",       "B002", "555-01", "Wing Spar Machining", "0010", 20.0, "CLOSE",  TECH_CCC, TECH_WC],

    # WO-200/0010 Charlie alone (quality inspector charge)
    ["WO-200", w(4, 4),  "Charlie Mendez","B003", "777-12", "Fuselage Drill",      "0010", 50.0, "CLOSE",  QA_CCC,   QA_WC],

    # WO-400/0010 Bob multi-week (incremental: 5 + 7 + 8 = 20 total)
    ["WO-400", w(3, 7),  "Bob Lee",       "B002", "555-02", "Sheet Metal Cut",     "0010", 5.0,  "ACTIVE", TECH_CCC, TECH_WC],
    ["WO-400", w(3, 14), "Bob Lee",       "B002", "555-02", "Sheet Metal Cut",     "0010", 7.0,  "ACTIVE", TECH_CCC, TECH_WC],
    ["WO-400", w(3, 21), "Bob Lee",       "B002", "555-02", "Sheet Metal Cut",     "0010", 8.0,  "CLOSE",  TECH_CCC, TECH_WC],

    # WO-500/0010 Bob alone (slow tech: actual > earned)
    ["WO-500", w(4, 11), "Bob Lee",       "B002", "555-01", "Wing Spar Machining", "0010", 30.0, "CLOSE",  TECH_CCC, TECH_WC],
]

# File B: covers parts 5*-* (overlaps 555-*) and 999-* (no earned coverage),
# plus Diana's new cleaning op (v1.7).
file_b_rows = [
    # WO-100/0010 -- duplicates of File A's rows; dedup should keep them once
    ["WO-100", w(2, 28), "Alice Johnson", "B001", "555-01", "Wing Spar Machining", "0010", 10.0, "ACTIVE", TECH_CCC, TECH_WC],
    ["WO-100", w(3, 7),  "Alice Johnson", "B001", "555-01", "Wing Spar Machining", "0010", 20.0, "CLOSE",  TECH_CCC, TECH_WC],
    ["WO-100", w(3, 7),  "Bob Lee",       "B002", "555-01", "Wing Spar Machining", "0010", 20.0, "CLOSE",  TECH_CCC, TECH_WC],

    # WO-100/0020 -- Alice charges, never closes (unmatched)
    ["WO-100", w(3, 7),  "Alice Johnson", "B001", "555-01", "Wing Spar Machining", "0020", 5.0,  "ACTIVE", TECH_CCC, TECH_WC],

    # WO-300/0030 -- entirely unmatched (no earned), part 999-05
    ["WO-300", w(3, 14), "Alice Johnson", "B001", "999-05", "Final Inspection",    "0030", 10.0, "ACTIVE", TECH_CCC, TECH_WC],
    ["WO-300", w(3, 14), "Charlie Mendez","B003", "999-05", "Final Inspection",    "0030", 8.0,  "ACTIVE", QA_CCC,   QA_WC],

    # WO-400/0010 again (duplicate of file A's last row)
    ["WO-400", w(3, 21), "Bob Lee",       "B002", "555-02", "Sheet Metal Cut",     "0010", 8.0,  "CLOSE",  TECH_CCC, TECH_WC],

    # WO-500/0010 again
    ["WO-500", w(4, 11), "Bob Lee",       "B002", "555-01", "Wing Spar Machining", "0010", 30.0, "CLOSE",  TECH_CCC, TECH_WC],

    # v1.7 -- WO-600/0010 Diana alone (cleaning crew, CCC=TECH but WC=WC-CLEAN)
    ["WO-600", w(4, 4),  "Diana Park",    "B004", "555-03", "Cleaning Bay 5",      "0010", 8.0,  "CLOSE",  CLEAN_CCC, CLEAN_WC],
]

make_xlsx("test-charges-555-777.xlsx", file_a_rows, "555-* and 777-* parts")
make_xlsx("test-charges-5x-and-999.xlsx", file_b_rows, "5*-* wildcard + 999-* parts + WO-600 cleaning")

# ----------------------------------------------------------------------------
# Hours Earned CSV
# ----------------------------------------------------------------------------
earned_lines = [
    "Order No,Year of Actual End Date,Month of Actual End Date,Oper No,Oper Type,Actual End Date,Hours Earned",
    "WO-100,2026,March,0010,MFG,3/7/2026,60",
    "WO-200,2026,April,0010,MFG,4/4/2026,40",
    "WO-400,2026,March,0010,MFG,3/21/2026,25",
    "WO-500,2026,April,0010,MFG,4/11/2026,25",
    "WO-600,2026,April,0010,MFG,4/4/2026,8",
    # Oper Type filter test: non-MFG row that all three consumers must skip.
    # Placed on WO-300/0080 (a QA op) -- WO-300 is the unmatched bucket, so this
    # row would not have produced any earned credit even WITHOUT the filter
    # (no charges to join against). That keeps PRD-VacBackfillTool §11.11 and
    # Score Card's per-employee CPI expectations unchanged while still
    # exercising the parser's filtering path and excluded-count counter.
    "WO-300,2026,March,0080,QA,3/15/2026,12",
    # Note: WO-100/0020 and WO-300/0030 deliberately absent -> unmatched test
]
(OUT / "test-hours-earned.csv").write_text("\n".join(earned_lines), encoding="utf-8")
print("Wrote test-hours-earned.csv")

# ----------------------------------------------------------------------------
# EV Details CSV (optional) -- exposes Program filter
# ----------------------------------------------------------------------------
ev_lines = [
    "Network,Order No,Part No,Plan Title,Program",
    "NET-A,WO-100,555-01,Wing Spar Machining,Aero-Alpha",
    "NET-A,WO-200,777-12,Fuselage Drill,Aero-Alpha",
    "NET-B,WO-300,999-05,Final Inspection,Aero-Beta",
    "NET-A,WO-400,555-02,Sheet Metal Cut,Aero-Alpha",
    "NET-A,WO-500,555-01,Wing Spar Machining,Aero-Alpha",
    "NET-A,WO-600,555-03,Cleaning Bay 5,Aero-Alpha",
]
(OUT / "test-ev-details.csv").write_text("\n".join(ev_lines), encoding="utf-8")
print("Wrote test-ev-details.csv")

# ----------------------------------------------------------------------------
# Expected results summary (for hand-verification)
# ----------------------------------------------------------------------------
print("\n=== Expected results (March 2026 + April 2026) ===")
print("Charges actuals (incremental):")
print("  Alice:   WO-100/0010 30h (Mar) + WO-100/0020 5h (Mar) + WO-300/0030 10h (Mar) = 45h")
print("  Bob:     WO-100/0010 20h (Mar) + WO-400/0010 5+7+8=20h (Mar) + WO-500/0010 30h (Apr) = 70h")
print("  Charlie: WO-200/0010 50h (Apr) + WO-300/0030 8h (Mar) = 58h")
print("  Diana:   WO-600/0010 8h (Apr) = 8h  (v1.7)")
print()
print("Earned (split + close month):")
print("  Alice:   WO-100/0010 60 x 30/50 = 36h (Mar) -> 36h earned")
print("  Bob:     WO-100/0010 60 x 20/50 = 24h (Mar) + WO-400/0010 25h (Mar) + WO-500/0010 25h (Apr) = 74h earned")
print("  Charlie: WO-200/0010 40h (Apr) -> 40h earned")
print("  Diana:   WO-600/0010 8h (Apr)  -> 8h  earned  (v1.7)")
print()
print("CPI (filtered, all months):")
print("  Alice:   36 / 45 = 0.80")
print("  Bob:     74 / 70 = 1.057")
print("  Charlie: 40 / 58 = 0.69")
print("  Diana:    8 /  8 = 1.00  (v1.7)")
print()
print("Split operations: WO-100/0010 (Alice 30/Bob 20)")
print("Unmatched charges: WO-100/0020 (5h, part 555-01), WO-300/0030 (18h, part 999-05)")
print("  Ranked: 999-05: 18h (top), 555-01: 5h")
print()
print("=== v1.7 CCC / Work Center filter expectations ===")
print("Default (no CCC/WC excluded):    4 techs visible (Alice, Bob, Charlie, Diana)")
print("Exclude CCC=QA:                  Charlie hidden -> 3 techs visible")
print("Exclude WC=WC-CLEAN:             Diana hidden   -> 3 techs visible")
print("Exclude CCC=QA AND WC=WC-CLEAN:  only Alice + Bob remain")
print("Manual click on Charlie's chip   -> Charlie back to ON regardless of CCC=QA filter")
print("  while CCC=QA is excluded:")

# ============================================================================
# v1.11 — Employee identity merge test (separate fixture trio, numeric badges)
#   Sam Carter:  00100200 (wk 6/6, older) + 100200 (wk 6/13, newer) -> zero-pad
#                auto-merge, canonical = 100200 (most recent). Program N.
#   Pat Lee:     300100 "Pat Lee" (contractor, wk 6/6) + 405000 "Pat A Lee"
#                (FTE, wk 6/13). Name-match candidate -> approve -> canonical
#                405000. CCC dots {CONTR, TECH}. Program P.
#   Jane Doe:    500500 (distinct baseline). Program N.
# ============================================================================
merge_rows = [
    ["WO-900", "6/6/2026 12:00:00 AM",  "Sam Carter", "00100200", "900-01", "Spar Assembly", "0010", 8.0,  "ACTIVE", "TECH",  "WC-MFG-1"],
    ["WO-900", "6/13/2026 12:00:00 AM", "Sam Carter", "100200",   "900-01", "Spar Assembly", "0010", 10.0, "CLOSE",  "TECH",  "WC-MFG-1"],
    ["WO-901", "6/6/2026 12:00:00 AM",  "Pat Lee",    "300100",   "901-01", "Panel Fit",     "0010", 12.0, "ACTIVE", "CONTR", "WC-MFG-2"],
    ["WO-901", "6/13/2026 12:00:00 AM", "Pat A Lee",  "405000",   "901-01", "Panel Fit",     "0010", 9.0,  "CLOSE",  "TECH",  "WC-MFG-2"],
    ["WO-902", "6/13/2026 12:00:00 AM", "Jane Doe",   "500500",   "902-01", "Final Assy",    "0010", 20.0, "CLOSE",  "TECH",  "WC-MFG-1"],
]
make_xlsx("merge-charges.xlsx", merge_rows, "identity-merge test")

merge_earned = [
    "Order No,Year of Actual End Date,Month of Actual End Date,Oper No,Oper Type,Actual End Date,Hours Earned",
    "WO-900,2026,June,0010,MFG,6/13/2026,15",
    "WO-901,2026,June,0010,MFG,6/13/2026,18",
    "WO-902,2026,June,0010,MFG,6/13/2026,16",
]
(OUT / "merge-hours-earned.csv").write_text("\n".join(merge_earned), encoding="utf-8")
print("Wrote merge-hours-earned.csv")

merge_ev = [
    "Network,Order No,Part No,Plan Title,Program",
    "NET-A,WO-900,900-01,Spar Assembly,N",
    "NET-A,WO-901,901-01,Panel Fit,P",
    "NET-A,WO-902,902-01,Final Assy,N",
]
(OUT / "merge-ev-details.csv").write_text("\n".join(merge_ev), encoding="utf-8")
print("Wrote merge-ev-details.csv")

print("\n=== v1.11 merge test expectations (eval wk 6/13/2026, ref month June=2 wks) ===")
print("Zero-pad auto: 00100200 + 100200 -> Sam Carter, canonical 100200; actual 18, earned 15")
print("Name match (approve): Pat Lee 300100 + Pat A Lee 405000 -> canonical 405000; actual 21, earned 18")
print("Jane Doe 500500 distinct (no merge). Programs: Sam=N, Pat=P, Jane=N")
